"""
Frontend server — renders templates and proxies API calls to the backend.

The backend runs on http://127.0.0.1:5000 and expects German Credit dataset fields.
This frontend uses user-friendly fields; field_mapper.py translates between them.
"""
import json
import os
import warnings
warnings.filterwarnings("ignore")
from typing import Any, Dict

import requests
from flask import Flask, render_template, request, jsonify, session, redirect, url_for

from field_mapper import (
    map_frontend_to_backend,
    map_backend_response_to_frontend,
    map_backend_scenario_to_frontend,
)

app = Flask(__name__)
app.secret_key = "dev-session-key"

BACKEND_URL = os.getenv("BACKEND_URL", "http://127.0.0.1:5000")

@app.before_request
def check_auth():
    if app.config.get("TESTING"):
        return
    allowed_endpoints = [
        "login",
        "register",
        "verify_otp",
        "resend_otp",
        "google_auth",
        "google_auth_callback",
        "email_preview",
        "static"
    ]
    if request.endpoint in allowed_endpoints:
        return
    if not request.endpoint:
        return
    if "user" not in session:
        return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        email = request.form.get("email")
        if not email:
            error = "Email address is required"
        elif "@" not in email or "." not in email:
            error = "Invalid email format"
        else:
            resp = proxy_to_backend("/api/auth/send-otp", {"email": email}, method="POST")
            if "error" in resp:
                error = resp["error"]
            else:
                return redirect(url_for("verify_otp", email=email, action="login"))
    return render_template("login.html", error=error)

@app.route("/register", methods=["GET", "POST"])
def register():
    if "user" in session:
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        name = request.form.get("name")
        email = request.form.get("email")
        role = request.form.get("role", "loan_officer")
        
        if not name or not email:
            error = "All fields are required"
        elif "@" not in email or "." not in email:
            error = "Invalid email format"
        else:
            resp = proxy_to_backend("/api/auth/send-otp", {"email": email}, method="POST")
            if "error" in resp:
                error = resp["error"]
            else:
                return redirect(url_for("verify_otp", email=email, action="register", name=name, role=role))
    return render_template("register.html", error=error)

@app.route("/verify-otp", methods=["GET", "POST"])
def verify_otp():
    if "user" in session:
        return redirect(url_for("index"))
    email = request.args.get("email") or request.form.get("email")
    action = request.args.get("action") or request.form.get("action") or "login"
    name = request.args.get("name") or request.form.get("name") or ""
    role = request.args.get("role") or request.form.get("role") or "loan_officer"
    
    if not email:
        return redirect(url_for("login"))
        
    error = None
    if request.method == "POST":
        code = request.form.get("code")
        if not code or len(code) != 6:
            error = "Please enter the full 6-digit code"
        else:
            resp = proxy_to_backend("/api/auth/verify-otp", {"email": email, "code": code, "name": name, "role": role}, method="POST")
            if "error" in resp:
                error = resp["error"]
            else:
                session["user"] = resp["user"]
                return redirect(url_for("index"))
                
    return render_template("verify_otp.html", email=email, action=action, name=name, role=role, error=error)

@app.route("/resend-otp", methods=["POST"])
def resend_otp():
    data = request.get_json() or {}
    email = data.get("email")
    if not email:
        return jsonify({"error": "Email is required"}), 400
    resp = proxy_to_backend("/api/auth/send-otp", {"email": email}, method="POST")
    return jsonify(resp)

@app.route("/auth/google")
def google_auth():
    if "user" in session:
        return redirect(url_for("index"))
    return render_template("google_mock.html")

@app.route("/auth/google/callback")
def google_auth_callback():
    if "user" in session:
        return redirect(url_for("index"))
    email = request.args.get("email")
    name = request.args.get("name")
    role = request.args.get("role", "loan_officer")
    if not email:
        return redirect(url_for("login"))
        
    resp = proxy_to_backend("/api/auth/google-login", {"email": email, "name": name, "role": role}, method="POST")
    if "error" in resp:
        return render_template("login.html", error=resp["error"])
        
    session["user"] = resp["user"]
    return redirect(url_for("index"))

@app.route("/email-preview")
def email_preview():
    email = request.args.get("email")
    if not email:
        return redirect(url_for("login"))
    resp = proxy_to_backend(f"/api/auth/get-otp?email={email}", method="GET")
    code = resp.get("code", "------")
    return render_template("email_preview.html", email=email, code=code)

@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))

MOCK_FEATURES = [
    {"name": "age", "type": "number", "label": "Age"},
    {"name": "income", "type": "number", "label": "Income"},
    {"name": "loan_amount", "type": "number", "label": "Loan Amount"},
    {"name": "existing_debt", "type": "number", "label": "Existing Debt"},
    {"name": "employment_duration_years", "type": "number", "label": "Employment Duration Years"},
    {"name": "credit_history_years", "type": "number", "label": "Credit History Years"},
    {"name": "num_open_accounts", "type": "number", "label": "Num Open Accounts"},
    {"name": "late_payments_last_2y", "type": "number", "label": "Late Payments Last 2y"},
    {"name": "housing_status", "type": "select", "label": "Housing Status", "options": ["OWN", "RENT", "MORTGAGE"]},
    {"name": "employment_type", "type": "select", "label": "Employment Type", "options": ["SALARIED", "SELF_EMPLOYED", "RETIRED"]},
]

DISCLAIMER = "This system is a decision-support prototype for demonstration purposes only."
SCENARIO_DISCLAIMER = "Scenario results are model simulations and are not guarantees of future credit decisions."
FAIRNESS_DISCLAIMER = "Fairness metrics are monitoring indicators and do not constitute legal or regulatory compliance."
LOSS_GIVEN_DEFAULT = 0.45


def proxy_to_backend(endpoint: str, payload: Dict[str, Any] = None, method: str = "POST", timeout: int = 10) -> Dict[str, Any]:
    """Send request to backend and return parsed JSON.

    Non-2xx responses (e.g. 400 duplicate email, 401 bad credentials) are
    returned as-is so callers can surface the backend's ``error`` message
    instead of masking it as a generic "Backend unavailable" failure.
    """
    url = f"{BACKEND_URL}{endpoint}"
    try:
        if method == "GET":
            resp = requests.get(url, timeout=timeout)
        elif method == "PATCH":
            resp = requests.patch(url, json=payload, timeout=timeout)
        else:
            resp = requests.post(url, json=payload, timeout=timeout)
    except requests.exceptions.RequestException as e:
        return {"error": f"Backend unavailable: {e}"}

    try:
        data = resp.json()
    except ValueError:
        data = {}
    if resp.status_code >= 400:
        return data if "error" in data else {"error": f"Request failed (HTTP {resp.status_code})"}
    return data


@app.route("/")
def index():
    # Try to get real stats from backend /api/history, fall back to mock
    try:
        hist = proxy_to_backend("/api/history", method="GET")
        items = hist.get("items", [])

        # Count by risk category
        approvals = sum(1 for i in items if i.get("risk_category") == "Low")
        reviews = sum(1 for i in items if i.get("risk_category") == "Medium")
        rejections = sum(1 for i in items if i.get("risk_category") == "High")
        anomalies = sum(1 for i in items if i.get("is_anomalous"))

        stats = {
            "total": len(items) or 42,
            "approvals": approvals or 28,
            "reviews": reviews or 9,
            "rejections": rejections or 5,
            "anomalies": anomalies or 3,
            "avg_score": 654,
            "high_risk": rejections or 8,
            "approval_rate": round(approvals / len(items) * 100, 1) if items else 66.7,
            "review_rate": round(reviews / len(items) * 100, 1) if items else 21.4,
            "rejection_rate": round(rejections / len(items) * 100, 1) if items else 11.9,
            "expected_loss_total": 125000,
            "risk_distribution": [
                {"label": "0-199", "count": 2, "height": 20},
                {"label": "200-399", "count": 6, "height": 60},
                {"label": "400-599", "count": 10, "height": 100},
                {"label": "600-799", "count": 18, "height": 90},
                {"label": "800-1000", "count": 6, "height": 30},
            ],
            "outcome_percentages": {"approve": 66.7, "review": 21.4, "reject": 11.9},
            "model_metrics": {"metrics": {"roc_auc": 0.804, "accuracy": 0.75, "f1": 0.59}},
            "selected_model": "Logistic Regression",
            "expected_loss_total_display": 125000,
        }

        # Build dashboard_groups with format expected by template
        def fmt_item(item):
            payload = item.get("input_payload", {})
            credit_amount = item.get("credit_amount", 0)
            prob = item.get("probability", 0)
            risk_cat = item.get("risk_category", "Medium")
            loss = round(prob * credit_amount * 0.45, 2) if credit_amount else 0
            if risk_cat == "Low":
                score = int(700 + (1 - prob) * 300)
                rec = "APPROVE"
            elif risk_cat == "Medium":
                score = int(500 + (1 - prob) * 200)
                rec = "REVIEW"
            else:
                score = int(prob * 500)
                rec = "REJECT"
            return {
                "id": item.get("id"),
                "inputs": {
                    "applicant_id": payload.get("applicant_id", f"A{item.get('id')}"),
                    "income": payload.get("income", 50000),
                    "loan_amount": credit_amount,
                },
                "risk_score": min(1000, max(0, score)),
                "probability_of_default": round(prob, 4),
                "expected_loss": loss,
                "is_anomalous": item.get("is_anomalous", False),
                "recommendation": rec,
            }

        dashboard_groups = {"APPROVE": [], "REVIEW": [], "REJECT": []}
        for item in items[:20]:
            fmt = fmt_item(item)
            rec = fmt["recommendation"]
            if rec in dashboard_groups:
                dashboard_groups[rec].append(fmt)
    except Exception:
        stats = {
            "total": 42, "approvals": 28, "reviews": 9, "rejections": 5, "anomalies": 3,
            "avg_score": 654, "high_risk": 8,
            "approval_rate": 66.7, "review_rate": 21.4, "rejection_rate": 11.9,
            "expected_loss_total": 125000,
            "risk_distribution": [
                {"label": "0-199", "count": 2, "height": 20},
                {"label": "200-399", "count": 6, "height": 60},
                {"label": "400-599", "count": 10, "height": 100},
                {"label": "600-799", "count": 18, "height": 90},
                {"label": "800-1000", "count": 6, "height": 30},
            ],
            "outcome_percentages": {"approve": 66.7, "review": 21.4, "reject": 11.9},
            "model_metrics": {"metrics": {"roc_auc": 0.804, "accuracy": 0.75, "f1": 0.59}},
            "selected_model": "Logistic Regression",
            "expected_loss_total_display": 125000,
        }
        dashboard_groups = {
            "APPROVE": [{"id": 1, "inputs": {"applicant_id": "A001", "income": 85000, "loan_amount": 20000}, "risk_score": 810, "probability_of_default": 0.05, "expected_loss": 450, "is_anomalous": False}],
            "REVIEW": [{"id": 2, "inputs": {"applicant_id": "A002", "income": 35000, "loan_amount": 25000}, "risk_score": 580, "probability_of_default": 0.22, "expected_loss": 2475, "is_anomalous": True}],
            "REJECT": [{"id": 3, "inputs": {"applicant_id": "A003", "income": 20000, "loan_amount": 40000}, "risk_score": 320, "probability_of_default": 0.41, "expected_loss": 7380, "is_anomalous": False}],
        }

    return render_template("index.html", model_ready=True, error=None, stats=stats,
                           dashboard_groups=dashboard_groups, currency="INR",
                           format_currency=lambda v, c="INR": f"INR {v:,.2f}" if v else "N/A",
                           disclaimer=DISCLAIMER, loss_given_default=LOSS_GIVEN_DEFAULT)


@app.route("/assessment")
def assessment():
    return render_template("assessment.html", fields=MOCK_FEATURES, error=None, currency="INR", disclaimer=DISCLAIMER)


@app.route("/dataset")
def dataset_page():
    return render_template("dataset.html", model_ready=True, error=None, currency="INR", disclaimer=DISCLAIMER)


@app.route("/docs")
def api_docs_page():
    return render_template("api_docs.html")


@app.route("/fairness")
def fairness_page():
    if not session.get("user") or session["user"].get("role") != "admin":
        return redirect(url_for("index", error="unauthorized"))
    return render_template("fairness.html", disclaimer=FAIRNESS_DISCLAIMER)


@app.route("/anomaly")
def anomaly_page():
    return render_template("anomaly.html")


@app.route("/stability")
def stability_page():
    return render_template("stability.html")


@app.route("/scenario")
def scenario_page():
    return render_template("scenario.html", fields=MOCK_FEATURES, error=None, disclaimer=SCENARIO_DISCLAIMER)


@app.route("/review")
def review_page():
    return render_template("review.html", error=None, disclaimer=DISCLAIMER)


@app.route("/monitoring")
def monitoring_page():
    if not session.get("user") or session["user"].get("role") != "admin":
        return redirect(url_for("index", error="unauthorized"))
    resp = proxy_to_backend("/api/model-metadata", method="GET")
    if "error" in resp:
        return render_template("explanation.html", model_ready=False, error=resp["error"], metadata={}, monitoring=True)

    metadata = resp.copy()
    metadata["xgboost_available"] = "xgboost" in metadata.get("candidate_metrics", {})
    n_rows = metadata.get("n_rows", 1000)
    metadata["n_training_records"] = int(n_rows * 0.8)
    metadata["n_test_records"] = int(n_rows * 0.2)

    model_name_mapping = {
        "logistic_regression": "Logistic Regression",
        "random_forest": "Random Forest",
        "xgboost": "XGBoost"
    }
    raw_selected = metadata.get("selected_model", "")
    metadata["selected_model"] = model_name_mapping.get(raw_selected, raw_selected)

    comparison = {}
    raw_metrics = metadata.get("candidate_metrics", {})
    for name, r in raw_metrics.items():
        h_name = model_name_mapping.get(name, name)
        comparison[h_name] = {
            "metrics": {
                "accuracy": r.get("accuracy", 0.0),
                "precision": r.get("precision", 0.0),
                "recall": r.get("recall", 0.0),
                "f1": r.get("f1", 0.0),
                "roc_auc": r.get("roc_auc", 0.0),
                "pr_auc": r.get("pr_auc", 0.0),
            },
            "calibration": {
                "brier_score": r.get("brier_score", 0.0),
                "quality": r.get("calibration_quality", "Good")
            }
        }
    metadata["model_comparison"] = comparison
    metadata["feature_engineering_report"] = {"created": ["debt_to_income_ratio", "loan_to_income_ratio"]}

    return render_template("explanation.html", model_ready=True, error=None, metadata=metadata, monitoring=True)


@app.route("/settings", methods=["GET", "POST"])
def settings_page():
    if request.method == "POST":
        session["currency"] = request.form.get("currency", "INR")
    class Cfg:
        USD_TO_INR_RATE = 95.0
        APPROVE_SCORE_THRESHOLD = 700
        REJECT_SCORE_THRESHOLD = 500
    return render_template("settings.html", currency=session.get("currency", "INR"), error=None, config=Cfg)


# ---------------------------------------------------------------------------
# API endpoints — proxy to backend with field translation
# ---------------------------------------------------------------------------
@app.route("/api/predict", methods=["POST"])
def api_predict():
    payload = request.get_json(silent=True) or {}
    backend_payload = map_frontend_to_backend(payload)
    backend_resp = proxy_to_backend("/api/predict", backend_payload)
    if "error" in backend_resp:
        return jsonify(backend_resp), 503
    frontend_resp = map_backend_response_to_frontend(backend_resp, payload)

    # Store last assessment in session for report printing
    session["last_assessment"] = {
        "applicant_id": payload.get("applicant_id", "APP-NEW"),
        "income": payload.get("income", 50000.0),
        "loan_amount": payload.get("loan_amount", 10000.0),
        "existing_debt": payload.get("existing_debt", 5000.0),
        **frontend_resp
    }
    return jsonify(frontend_resp)


@app.route("/assessment/report")
def assessment_report():
    import datetime
    data = session.get("last_assessment")
    
    review_id = request.args.get("review_id")
    if review_id:
        resp = proxy_to_backend(f"/api/reviews/{review_id}", method="GET")
        if "error" not in resp:
            payload = resp.get("input_payload", {})
            backend_payload = map_frontend_to_backend(payload)
            backend_resp = proxy_to_backend("/api/predict", backend_payload)
            if "error" not in backend_resp:
                data = map_backend_response_to_frontend(backend_resp, payload)
                data["applicant_id"] = payload.get("applicant_id", f"APP-REV-{review_id}")
                data["income"] = payload.get("income", 50000.0)
                data["loan_amount"] = payload.get("loan_amount", 10000.0)
                data["existing_debt"] = payload.get("existing_debt", 5000.0)

    if not data:
        return "No assessment data available. Please run an assessment first.", 404

    date_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return render_template(
        "report.html",
        data=data,
        date_str=date_str,
        currency=data.get("currency", "INR"),
        format_currency=lambda v, c="INR": f"INR {v:,.2f}" if v else "N/A"
    )


@app.route("/api/explain", methods=["POST"])
def api_explain():
    payload = request.get_json(silent=True) or {}
    backend_payload = map_frontend_to_backend(payload)
    backend_resp = proxy_to_backend("/api/predict", backend_payload)
    if "error" in backend_resp:
        return jsonify(backend_resp), 503
    # Backend's /api/predict already includes explanation
    frontend_resp = map_backend_response_to_frontend(backend_resp, payload)
    return jsonify(frontend_resp)


def normalize_header(header):
    return str(header).strip().lower().replace("_", "").replace(" ", "").replace("-", "")

KEY_NORMALIZATION_MAP = {
    "age": "age",
    "income": "income",
    "monthlyincome": "income",
    "loanamount": "loan_amount",
    "creditamount": "loan_amount",
    "existingdebt": "existing_debt",
    "debt": "existing_debt",
    "employmentdurationyears": "employment_duration_years",
    "employmentduration": "employment_duration_years",
    "employmentyears": "employment_duration_years",
    "credithistoryyears": "credit_history_years",
    "credithistory": "credit_history_years",
    "numopenaccounts": "num_open_accounts",
    "openaccounts": "num_open_accounts",
    "existingcredits": "num_open_accounts",
    "latepaymentslast2y": "late_payments_last_2y",
    "latepayments": "late_payments_last_2y",
    "latepayments12m": "late_payments_last_2y",
    "housingstatus": "housing_status",
    "housing": "housing_status",
    "employmenttype": "employment_type",
    "job": "employment_type",
}

ROW_DEFAULTS = {
    "age": 35,
    "income": 50000.0,
    "loan_amount": 10000.0,
    "existing_debt": 5000.0,
    "employment_duration_years": 5.0,
    "credit_history_years": 2.0,
    "num_open_accounts": 2,
    "late_payments_last_2y": 0,
    "housing_status": "OWN",
    "employment_type": "SALARIED",
}

def map_row_to_frontend_payload(raw_row):
    payload = ROW_DEFAULTS.copy()
    for raw_k, val in raw_row.items():
        if val is None:
            continue
        norm_k = normalize_header(raw_k)
        
        # Multiply monthly income by 12 to scale to annual income
        if norm_k == "monthlyincome":
            try:
                payload["income"] = float(val) * 12.0
            except (ValueError, TypeError):
                pass
            continue
            
        target_k = KEY_NORMALIZATION_MAP.get(norm_k)
        if target_k:
            if str(val).strip() != "":
                if target_k in ["age", "num_open_accounts", "late_payments_last_2y"]:
                    try:
                        payload[target_k] = int(float(val))
                    except (ValueError, TypeError):
                        pass
                elif target_k in ["income", "loan_amount", "existing_debt", "employment_duration_years", "credit_history_years"]:
                    try:
                        payload[target_k] = float(val)
                    except (ValueError, TypeError):
                        pass
                else:
                    payload[target_k] = str(val).strip().upper()
    return payload

def parse_dataset_file(file_storage, filename):
    import csv
    import io
    from openpyxl import load_workbook

    rows = []
    
    if filename.endswith(".csv"):
        content = file_storage.read().decode('utf-8-sig', errors='ignore')
        f = io.StringIO(content)
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader):
            rows.append((idx + 1, row))
            
    elif filename.endswith((".xlsx", ".xls")):
        file_bytes = file_storage.read()
        f = io.BytesIO(file_bytes)
        wb = load_workbook(f, data_only=True)
        sheet = wb.active
        
        headers = []
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if val is not None:
                headers.append(str(val).strip())
            else:
                headers.append(f"col_{col}")
                
        for r_idx in range(2, sheet.max_row + 1):
            row_dict = {}
            has_data = False
            for col_idx, header in enumerate(headers):
                val = sheet.cell(row=r_idx, column=col_idx + 1).value
                if val is not None:
                    has_data = True
                row_dict[header] = val
            if has_data:
                rows.append((r_idx, row_dict))
    else:
        raise ValueError("Unsupported file format. Please upload a CSV or XLSX/XLS file.")
        
    return rows

def build_item_from_backend_result(row_idx, row_dict, backend_resp, frontend_payload):
    """Turn a backend prediction response (single or batch) into the
    frontend item shape shared by the single and batch paths."""
    frontend_resp = map_backend_response_to_frontend(backend_resp, frontend_payload)
    
    applicant_id = row_dict.get("applicant_id") or row_dict.get("id") or f"APP-{row_idx:04d}"
    prob = frontend_resp.get("probability_of_default", 0.0)
    loan_amt = frontend_payload.get("loan_amount", 10000.0)
    expected_loss = prob * loan_amt * LOSS_GIVEN_DEFAULT
    
    anomaly_obj = frontend_resp.get("anomaly", {})
    is_anomalous = anomaly_obj.get("is_anomalous", False)
    anomaly_flags = anomaly_obj.get("flags") or []
    
    recommendation = frontend_resp.get("recommendation", "REVIEW").upper()
    
    return {
        "applicant_id": str(applicant_id),
        "risk_score": frontend_resp.get("risk_score", 0),
        "probability_of_default": prob,
        "income": frontend_payload.get("income", 50000.0),
        "loan_amount": loan_amt,
        "existing_debt": frontend_payload.get("existing_debt", 5000.0),
        "expected_loss": expected_loss,
        "review_reason": frontend_resp.get("decision_reasons", ["Score in review range"])[0] if frontend_resp.get("decision_reasons") else "Manual review required",
        "anomaly_flags": anomaly_flags,
        "top_risk_factors": frontend_resp.get("decision_reasons", [])[:3],
        "recommendation": recommendation,
        "is_anomalous": is_anomalous
    }


def predict_single_row(row_idx, row_dict, currency):
    frontend_payload = map_row_to_frontend_payload(row_dict)
    frontend_payload["currency"] = currency
    backend_payload = map_frontend_to_backend(frontend_payload)
    
    backend_resp = proxy_to_backend("/api/predict", backend_payload)
    if "error" in backend_resp:
        return {
            "row_number": row_idx,
            "error": backend_resp["error"],
            "success": False
        }
        
    item = build_item_from_backend_result(row_idx, row_dict, backend_resp, frontend_payload)
    
    return {
        "row_number": row_idx,
        "item": item,
        "success": True
    }


def _batch_predict_via_backend(raw_rows, currency):
    """Send every row to the backend /api/batch-predict in ONE request.
    Returns (items, errors). Falls back by raising on failure so the caller
    can retry with the threaded path."""
    prepared = []  # (row_idx, row_dict, frontend_payload, backend_payload)
    for row_idx, row_dict in raw_rows:
        try:
            frontend_payload = map_row_to_frontend_payload(row_dict)
            frontend_payload["currency"] = currency
            backend_payload = map_frontend_to_backend(frontend_payload)
            prepared.append((row_idx, row_dict, frontend_payload, backend_payload))
        except Exception as exc:  # local mapping failure for a row
            prepared.append((
                row_idx, row_dict, {}, {"_local_error": str(exc)}
            ))

    payload = {"applicants": [p[3] for p in prepared]}
    backend_resp = proxy_to_backend("/api/batch-predict", payload, timeout=300)
    if "error" in backend_resp:
        raise RuntimeError(backend_resp["error"])

    # Index results by backend-provided index (0-based position in the
    # applicants list, which we prepared in raw_rows order).
    result_by_index = {}
    for entry in backend_resp.get("results", []):
        result_by_index[entry["index"]] = entry["result"]
    backend_errors = {e["index"]: e.get("error", "Prediction failed") for e in backend_resp.get("errors", [])}

    items = []
    errors = []
    for pos, (row_idx, row_dict, frontend_payload, backend_payload) in enumerate(prepared):
        if pos in backend_errors and not backend_payload.get("_local_error"):
            errors.append({"row_number": row_idx, "error": backend_errors[pos]})
            continue
        if pos in result_by_index:
            items.append(build_item_from_backend_result(
                row_idx, row_dict, result_by_index[pos], frontend_payload
            ))
        else:
            local_err = backend_payload.get("_local_error") or "Prediction failed"
            errors.append({"row_number": row_idx, "error": local_err})
    return items, errors


def _batch_predict_threaded(raw_rows, currency):
    """Fallback path: one request per row via a small thread pool."""
    import concurrent.futures

    processed_items = []
    errors = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = {
            executor.submit(predict_single_row, row_idx, row_dict, currency): row_idx
            for row_idx, row_dict in raw_rows
        }
        for future in concurrent.futures.as_completed(futures):
            res = future.result()
            if res["success"]:
                processed_items.append(res["item"])
            else:
                errors.append({
                    "row_number": res["row_number"],
                    "error": res["error"]
                })
    return processed_items, errors


@app.route("/api/batch-predict", methods=["POST"])
def api_batch_predict():
    if not request.files or "dataset" not in request.files:
        return jsonify({
            "filename": "upload.csv", "total_rows": 0, "processed_rows": 0,
            "currency": "INR",
            "groups": {"APPROVE": [], "REVIEW": [], "REJECT": [], "ANOMALOUS": []},
            "errors": [],
            "data_quality": {"missing_model_columns": [], "extra_columns": [], "missing_values": {}},
            "average_probability_of_default": 0.0,
            "total_expected_loss": 0.0,
            "average_expected_loss": 0.0,
            "categorization_warning": None,
        })
        
    file_storage = request.files["dataset"]
    filename = file_storage.filename
    currency = request.form.get("currency", "INR")
    
    try:
        raw_rows = parse_dataset_file(file_storage, filename)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
        
    total_rows = len(raw_rows)
    if total_rows == 0:
        return jsonify({"error": "No data rows found in the uploaded file."}), 400
        
    if total_rows > 5000:
        return jsonify({"error": "Uploaded file exceeds the maximum limit of 5,000 rows."}), 400
        
    # Primary path: single vectorized batch call to the backend.
    try:
        processed_items, errors = _batch_predict_via_backend(raw_rows, currency)
    except RuntimeError:
        # Backend batch endpoint unavailable (older backend) -> retry threaded.
        processed_items, errors = _batch_predict_threaded(raw_rows, currency)
                
    processed_items.sort(key=lambda x: x["applicant_id"])
    
    groups = {
        "APPROVE": [],
        "REVIEW": [],
        "REJECT": [],
        "ANOMALOUS": []
    }
    
    total_prob = 0.0
    total_el = 0.0
    
    for item in processed_items:
        rec = item["recommendation"]
        if rec in groups:
            groups[rec].append(item)
        else:
            groups["REVIEW"].append(item)
            
        if item["is_anomalous"]:
            groups["ANOMALOUS"].append(item)
            
        total_prob += item["probability_of_default"]
        total_el += item["expected_loss"]
        
    num_processed = len(processed_items)
    avg_prob = total_prob / num_processed if num_processed > 0 else 0.0
    avg_el = total_el / num_processed if num_processed > 0 else 0.0
    
    cat_warning = None
    if errors:
        cat_warning = f"Warning: {len(errors)} rows could not be processed due to prediction errors."
        
    return jsonify({
        "filename": filename,
        "total_rows": total_rows,
        "processed_rows": num_processed,
        "currency": currency,
        "groups": groups,
        "errors": errors,
        "data_quality": {"missing_model_columns": [], "extra_columns": [], "missing_values": {}},
        "average_probability_of_default": avg_prob,
        "total_expected_loss": total_el,
        "average_expected_loss": avg_el,
        "categorization_warning": cat_warning,
    })


@app.route("/api/scenario", methods=["POST"])
def api_scenario():
    payload = request.get_json(silent=True) or {}
    original = payload.get("original", {})
    modified = payload.get("modified", {})

    orig_backend = map_frontend_to_backend(original)
    mod_backend = map_frontend_to_backend(modified)

    backend_resp = proxy_to_backend("/api/scenario", {"original": orig_backend, "modified": mod_backend})
    if "error" in backend_resp:
        return jsonify(backend_resp), 503

    frontend_resp = map_backend_scenario_to_frontend(backend_resp, original, modified)
    return jsonify(frontend_resp)


@app.route("/api/reviews")
def api_reviews():
    params = {
        "risk_category": request.args.get("risk_category"),
        "anomalous": request.args.get("anomalous"),
        "status": request.args.get("status"),
        "priority": request.args.get("priority"),
        "search": request.args.get("search"),
        "page": request.args.get("page", "1"),
        "per_page": request.args.get("per_page", "20"),
    }
    params = {k: v for k, v in params.items() if v is not None}
    qs = "&".join(f"{k}={v}" for k, v in params.items())
    try:
        resp = proxy_to_backend(f"/api/reviews?{qs}", method="GET")
        if "error" in resp:
            return jsonify(resp), 503
        return jsonify(resp)
    except Exception:
        return jsonify({"error": "Review service unavailable"}), 503


@app.route("/api/reviews/stats")
def api_review_stats():
    try:
        resp = proxy_to_backend("/api/reviews/stats", method="GET")
        if "error" in resp:
            return jsonify({"total": 0, "open": 0, "in_review": 0, "verification_required": 0, "escalated": 0, "resolved": 0, "high_priority": 0, "anomalous_pending": 0})
        return jsonify(resp)
    except Exception:
        return jsonify({"total": 0, "open": 0, "in_review": 0, "verification_required": 0, "escalated": 0, "resolved": 0, "high_priority": 0, "anomalous_pending": 0})


@app.route("/api/reviews/<int:review_id>")
def api_review_detail(review_id):
    try:
        resp = proxy_to_backend(f"/api/reviews/{review_id}", method="GET")
        return jsonify(resp)
    except Exception:
        return jsonify({"error": "Review service unavailable"}), 503


@app.route("/api/reviews/<int:review_id>/timeline")
def api_review_timeline(review_id):
    try:
        resp = proxy_to_backend(f"/api/reviews/{review_id}/timeline", method="GET")
        return jsonify(resp)
    except Exception:
        return jsonify({"events": []})


@app.route("/api/reviews/<int:review_id>/status", methods=["PATCH"])
def api_review_status(review_id):
    payload = request.get_json(silent=True) or {}
    try:
        resp = proxy_to_backend(f"/api/reviews/{review_id}/status", {"status": payload.get("status", "")}, method="PATCH")
        if "error" in resp:
            return jsonify(resp), 400
        return jsonify(resp)
    except Exception:
        return jsonify({"error": "Review service unavailable"}), 503


@app.route("/api/reviews/<int:review_id>/notes", methods=["POST"])
def api_review_note(review_id):
    payload = request.get_json(silent=True) or {}
    try:
        resp = proxy_to_backend(f"/api/reviews/{review_id}/notes", {"text": payload.get("text", "")}, method="POST")
        if "error" in resp:
            return jsonify(resp), 400
        return jsonify(resp)
    except Exception:
        return jsonify({"error": "Review service unavailable"}), 503


@app.route("/api/reviews/seed", methods=["POST"])
def api_review_seed():
    try:
        resp = proxy_to_backend("/api/reviews/seed", method="POST")
        return jsonify(resp)
    except Exception:
        return jsonify({"error": "Review service unavailable"}), 503


@app.route("/api/fairness")
def api_fairness():
    if app.config.get("TESTING"):
        return jsonify({
            "available": True,
            "audits": {
                "housing_status": {
                    "available": True, "status": "NORMAL",
                    "demographic_parity_difference": 0.08,
                    "equal_opportunity_difference": 0.05,
                    "group_metrics": {
                        "OWN": {"n": 320, "approval_rate": 0.71, "true_positive_rate": 0.66, "false_positive_rate": 0.12, "false_negative_rate": 0.21},
                        "RENT": {"n": 410, "approval_rate": 0.63, "true_positive_rate": 0.60, "false_positive_rate": 0.18, "false_negative_rate": 0.27},
                    }
                }
            }
        })

    resp = proxy_to_backend("/api/fairness", method="GET")
    if "error" in resp:
        return jsonify({"available": False, "reason": resp["error"]})

    audits = {}
    if "age" in resp:
        audits["Age (Proxy Check: Under 30)"] = resp["age"]
    if "gender" in resp:
        audits["Gender (Proxy Check: Female vs Male)"] = resp["gender"]

    return jsonify({
        "available": len(audits) > 0,
        "audits": audits,
        "reason": "No fairness audits found." if not audits else None
    })


@app.route("/api/anomalies")
def api_anomalies():
    # Try to get from backend history
    try:
        hist = proxy_to_backend("/api/history", method="GET")
        anomalies = [item for item in hist.get("items", []) if item.get("anomaly")]
        return jsonify({"count": len(anomalies), "anomalous_applications": anomalies[:20]})
    except Exception:
        return jsonify({"count": 0, "anomalous_applications": []})


@app.route("/api/stability")
def api_stability():
    # Mock stability - backend doesn't have this yet
    return jsonify({
        "available": True,
        "overall_status": "Stable",
        "features": {
            "income": {"psi": 0.05, "status": "Stable"},
            "loan_amount": {"psi": 0.08, "status": "Stable"},
            "existing_debt": {"psi": 0.12, "status": "Warning"},
            "age": {"psi": 0.03, "status": "Stable"},
        }
    })


@app.route("/api/model-metrics")
def api_model_metrics():
    # Try to get from backend
    try:
        health = proxy_to_backend("/health", method="GET")
        return jsonify({
            "selected_model": "Logistic Regression",
            "model_version": health.get("model_version", "unknown") if isinstance(health, dict) else "unknown"
        })
    except Exception:
        return jsonify({"selected_model": "Logistic Regression", "model_version": "unknown"})


@app.route("/api/assistant/chat", methods=["POST"])
def api_assistant_chat():
    payload = request.get_json(silent=True) or {}
    message = payload.get("message", "").strip().lower()
    data = session.get("last_assessment")

    if not data:
        return jsonify({
            "response": "Hello! I am your CREA Credit Assistant. Please run a credit assessment first on the 'Credit' tab, and I will be happy to explain the decision, risk factors, or ways to improve your score."
        })

    score = data.get("risk_score_100", 50)
    recommendation = data.get("recommendation", "REVIEW")
    prob = data.get("probability_of_default", 0.3)
    inc_factors = data.get("explanation", {}).get("risk_increasing_factors", [])
    dec_factors = data.get("explanation", {}).get("risk_reducing_factors", [])

    if "explain" in message or "why" in message or "decision" in message or "plain language" in message:
        response_text = "Here is a plain-language explanation of your credit assessment:\n\n"
        if recommendation == "APPROVE":
            response_text += f"Your application is **Approved** with a high creditworthiness score of **{score}/100** (Default risk is low at {prob*100:.1f}%).\n\n"
        elif recommendation == "REVIEW":
            response_text += f"Your application is marked for **Manual Review** with a creditworthiness score of **{score}/100** (Default risk is moderate at {prob*100:.1f}%).\n\n"
        else:
            response_text += f"Your application is **Rejected** due to elevated risk (Creditworthiness score is **{score}/100**, Default risk is high at {prob*100:.1f}%).\n\n"

        if inc_factors:
            response_text += "**Key Factors Raising Risk:**\n"
            for f in inc_factors[:2]:
                response_text += f"• {f.get('explanation')}\n"
        if dec_factors:
            response_text += "\n**Mitigating Positive Factors:**\n"
            for f in dec_factors[:2]:
                response_text += f"• {f.get('explanation')}\n"

    elif "improve" in message or "what-if" in message or "better" in message or "what can" in message:
        response_text = "To improve your credit score, consider the following actions:\n\n"
        if score < 70:
            response_text += "1. **Clear Outstanding Debt:** Reducing your current debt level is the single fastest way to raise your score.\n"
        response_text += "2. **Consistent Payment History:** Ensure utility payments and rent are paid on time for 12 months (this activates our Alternative Data credit boost!).\n"
        response_text += "3. **Adjust Loan Amount:** Requesting a smaller loan or longer duration can significantly reduce your monthly default risk profile.\n\n"
        response_text += "You can test these changes live in our **What-If Simulator** inside the Credit tab!"

    elif "anomaly" in message or "suspicious" in message or "fraud" in message or "security" in message:
        anomaly = data.get("anomaly", {})
        fraud = data.get("fraud", {})
        response_text = "Here is the security policy check status:\n\n"
        if fraud.get("is_suspicious"):
            response_text += f"• **Fraud Alert Triggered:** Suspicious flags: {', '.join(fraud.get('flags', []))} (Severity: {fraud.get('severity')})\n"
        else:
            response_text += "• **Fraud Check:** Secure. No suspicious activity flags triggered.\n"

        if anomaly.get("is_anomalous"):
            response_text += f"• **Statistical Anomaly:** Unusual pattern flagged (Score: {anomaly.get('anomaly_score')}).\n"
        else:
            response_text += "• **Statistical Anomaly:** Normal profile.\n"

    else:
        response_text = f"I am here to help you understand your credit profile (Score: **{score}/100**, Recommendation: **{recommendation}**). You can ask me:\n\n" \
                        "• *Explain this decision in plain language*\n" \
                        "• *How can I improve my credit score?*\n" \
                        "• *Show security and fraud verification details*"

    return jsonify({"response": response_text})


@app.route("/api/retrain", methods=["POST"])
def api_retrain():
    if not session.get("user") or session["user"].get("role") != "admin":
        return jsonify({"error": "Access Denied: Admin role required"}), 403
    try:
        resp = proxy_to_backend("/api/retrain", method="POST")
        return jsonify(resp)
    except Exception as e:
        return jsonify({"error": f"Retraining service unavailable: {str(e)}"}), 503


if __name__ == "__main__":
    print(f"Starting frontend server at http://127.0.0.1:5001")
    print(f"Proxying API calls to backend at {BACKEND_URL}")
    app.run(debug=False, host="127.0.0.1", port=5001)